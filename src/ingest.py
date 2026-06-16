"""Ingestion and normalization (spec FR-1/FR-2, task T4).

Reads the eight domain-list sheets plus ``EU Web Site URLs`` from the source
workbook, dynamically locating header rows, normalizes domains to their
registrable (eTLD+1) form, de-duplicates across sheets, and merges the EU seed
fields onto matching domains.
"""

from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
import tldextract

from .models import InputDomain

log = logging.getLogger(__name__)

# The eight domain-list sheets (regional + subsidiary + non-classified).
DOMAIN_SHEETS = [
    "North America",
    "EMEA",
    "LATAM",
    "APJ",
    "NA - Public Sector",
    "Hyve",
    "Shyft",
    "Non-Classified",
]
EU_SHEET = "EU Web Site URLs"

DEFAULT_SOURCE = "url-regional-class.xlsx"

# Offline-capable extractor: use the snapshot bundled with tldextract so the
# pipeline does not depend on fetching the public-suffix list at import time.
# ``include_psl_private_domains`` so private suffixes (e.g. CentralNic's
# ``br.com``) are honored, keeping e.g. ``synnex.br.com`` and
# ``synnexcorp.br.com`` as distinct registrable domains.
_extract = tldextract.TLDExtract(
    suffix_list_urls=(), include_psl_private_domains=True
)


def registrable_domain(value: str) -> str:
    """Normalize a raw cell/URL value to its registrable domain (eTLD+1).

    Lowercases, strips whitespace, drops scheme/path/query and a leading
    ``www.``, then derives eTLD+1. Falls back to the cleaned host if the suffix
    is unknown.
    """
    if not value:
        return ""
    s = str(value).strip().lower()
    if not s:
        return ""
    # Drop scheme.
    s = re.sub(r"^[a-z][a-z0-9+.\-]*://", "", s)
    # Drop everything after the first slash, ?, #, : (port), or whitespace.
    s = re.split(r"[/?#\s]", s, maxsplit=1)[0]
    s = s.split(":")[0]
    if s.startswith("www."):
        s = s[4:]
    s = s.strip(".")
    if not s:
        return ""
    ext = _extract(s)
    if ext.domain and ext.suffix:
        return f"{ext.domain}.{ext.suffix}"
    # Unknown suffix (or bare label): return the cleaned host as-is.
    return s


def _clean(value) -> str:
    return str(value).strip() if value is not None else ""


def _find_header_row(ws, expected_first: str, max_scan: int = 10) -> int:
    """Return the 1-based row index whose first cell equals ``expected_first``."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        first = _clean(row[0]).lower() if row else ""
        if first == expected_first.lower():
            return i
    # Fallback: assume row 3 for formatted sheets.
    return 3


def _read_domain_sheet(ws) -> list[InputDomain]:
    """Read one 7-column domain-list sheet into InputDomain rows."""
    header_row = _find_header_row(ws, "Domain")
    rows: list[InputDomain] = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        original = _clean(row[0])
        if not original:
            continue
        reg = registrable_domain(original)
        if not reg:
            continue
        rows.append(
            InputDomain(
                domain=reg,
                original=original,
                tld=_clean(row[1]) if len(row) > 1 else "",
                country=_clean(row[2]) if len(row) > 2 else "",
                source=_clean(row[3]) if len(row) > 3 else "",
                hosting=_clean(row[4]) if len(row) > 4 else "",
                region=_clean(row[5]) if len(row) > 5 else "",
                classification_method=_clean(row[6]) if len(row) > 6 else "",
            )
        )
    return rows


def _read_eu_seed(ws) -> dict[str, dict]:
    """Read the EU Web Site URLs sheet into {registrable_domain: seed_fields}."""
    header_row = _find_header_row(ws, "URL")
    seeds: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row:
            continue
        url = _clean(row[0])
        if not url:
            continue
        # Prefer the explicit Root/Base Domain columns; fall back to the URL.
        root = _clean(row[2]) if len(row) > 2 else ""
        base = _clean(row[1]) if len(row) > 1 else ""
        reg = registrable_domain(root or base or url)
        if not reg:
            continue
        seed = {
            "seed_url": url,
            "seed_site_status": _clean(row[5]) if len(row) > 5 else "",
            "seed_purpose": _clean(row[6]) if len(row) > 6 else "",
            "seed_cms": _clean(row[7]) if len(row) > 7 else "",
            "seed_hosting_provider": _clean(row[8]) if len(row) > 8 else "",
        }
        # Keep the first non-empty seed per domain.
        seeds.setdefault(reg, seed)
    return seeds


def load_domains(
    xlsx_path: str | Path = DEFAULT_SOURCE,
    *,
    sheets: Optional[Iterable[str]] = None,
) -> list[InputDomain]:
    """Load, normalize, de-duplicate, and seed-merge all source domains.

    Returns a list of unique InputDomain objects keyed by registrable domain.
    De-duplication keeps the first occurrence (sheet order in ``DOMAIN_SHEETS``)
    but back-fills empty fields from later occurrences.
    """
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    target_sheets = list(sheets) if sheets is not None else DOMAIN_SHEETS

    by_domain: dict[str, InputDomain] = {}
    per_sheet_counts: dict[str, int] = {}
    total_rows = 0

    for name in target_sheets:
        if name not in wb.sheetnames:
            log.warning("sheet %r not found; skipping", name)
            continue
        sheet_rows = _read_domain_sheet(wb[name])
        per_sheet_counts[name] = len(sheet_rows)
        total_rows += len(sheet_rows)
        for item in sheet_rows:
            if item.domain not in by_domain:
                by_domain[item.domain] = item
            else:
                # Back-fill any empty fields from this duplicate.
                existing = by_domain[item.domain]
                for f in ("tld", "country", "source", "hosting", "region",
                          "classification_method"):
                    if not getattr(existing, f) and getattr(item, f):
                        setattr(existing, f, getattr(item, f))

    # Merge EU seed fields.
    seeded = 0
    if EU_SHEET in wb.sheetnames:
        seeds = _read_eu_seed(wb[EU_SHEET])
        for reg, seed in seeds.items():
            if reg in by_domain:
                obj = by_domain[reg]
                obj.seed_url = seed["seed_url"] or None
                obj.seed_site_status = seed["seed_site_status"] or None
                obj.seed_purpose = seed["seed_purpose"] or None
                obj.seed_cms = seed["seed_cms"] or None
                obj.seed_hosting_provider = seed["seed_hosting_provider"] or None
                seeded += 1

    wb.close()

    result = list(by_domain.values())
    dropped = total_rows - len(result)
    log.info(
        "ingest: %d raw rows across %d sheets -> %d unique domains "
        "(%d cross-sheet duplicates collapsed); %d EU-seeded",
        total_rows, len(per_sheet_counts), len(result), dropped, seeded,
    )
    for name, cnt in per_sheet_counts.items():
        log.info("  %-22s %d", name, cnt)

    return result


def load_eu_seed(xlsx_path: str | Path = DEFAULT_SOURCE) -> dict[str, dict]:
    """Public helper: return the EU seed map (used by validation)."""
    xlsx_path = Path(xlsx_path)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if EU_SHEET not in wb.sheetnames:
            return {}
        return _read_eu_seed(wb[EU_SHEET])
    finally:
        wb.close()
