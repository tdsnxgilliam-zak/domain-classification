"""Output writers (spec Section 6, task T12).

Writes the master CSV and the deliverable xlsx (Inventory + Summary sheets). The
source workbook is never opened for writing.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import ClassifiedDomain, Summary

log = logging.getLogger(__name__)

# The 12 spec fields, in the exact required order, followed by provenance.
SPEC_HEADERS = [
    "Domain",
    "Website behind URL?",
    "Redirect?",
    "Redirect target",
    "Page opens / renders?",
    "Email functionality?",
    "Date of last update found",
    "Content status",
    "Likely purpose",
    "Defensive-registration likelihood",
    "Release recommendation",
    "Evidence / notes",
]
PROVENANCE_HEADERS = ["Region", "Source", "Country"]
ALL_HEADERS = SPEC_HEADERS + PROVENANCE_HEADERS


def _row_values(r: ClassifiedDomain) -> list:
    return [
        r.domain,
        r.website_behind_url,
        r.redirect,
        r.redirect_target,
        r.page_opens_renders,
        r.email_functionality,
        r.date_last_update,
        r.content_status,
        r.likely_purpose,
        r.defensive_likelihood,
        r.release_recommendation,
        r.evidence_notes,
        r.region,
        r.source,
        r.country,
    ]


def write_csv(rows: list[ClassifiedDomain], path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow(ALL_HEADERS)
        for r in rows:
            writer.writerow(_row_values(r))
    log.info("wrote CSV: %s (%d rows)", path, len(rows))
    return path


# --------------------------------------------------------------------------- #
# xlsx
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13)
LABEL_FONT = Font(bold=True)

REC_FILLS = {
    "Keep": PatternFill("solid", fgColor="C6EFCE"),
    "Review": PatternFill("solid", fgColor="FFEB9C"),
    "Candidate for release": PatternFill("solid", fgColor="FCE4D6"),
    "Do not release": PatternFill("solid", fgColor="F8CBAD"),
    "Unknown": PatternFill("solid", fgColor="D9D9D9"),
}

COL_WIDTHS = {
    "Domain": 28, "Website behind URL?": 16, "Redirect?": 11,
    "Redirect target": 40, "Page opens / renders?": 18,
    "Email functionality?": 17, "Date of last update found": 20,
    "Content status": 15, "Likely purpose": 50,
    "Defensive-registration likelihood": 18, "Release recommendation": 20,
    "Evidence / notes": 70, "Region": 16, "Source": 16, "Country": 16,
}


def _write_inventory_sheet(ws, rows: list[ClassifiedDomain]) -> None:
    ws.title = "Inventory"
    ws.append(ALL_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    rec_col = ALL_HEADERS.index("Release recommendation") + 1
    for r in rows:
        ws.append(_row_values(r))
        rec_cell = ws.cell(row=ws.max_row, column=rec_col)
        fill = REC_FILLS.get(r.release_recommendation)
        if fill:
            rec_cell.fill = fill

    # Widths.
    for idx, header in enumerate(ALL_HEADERS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(header, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(ALL_HEADERS))}{ws.max_row}"


def _write_summary_sheet(ws, summary: Summary) -> None:
    ws.title = "Summary"
    ws["A1"] = "Worldwide Domain Inventory - Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    metrics = [
        ("Total domains reviewed", summary.total),
        ("Recommended: Keep", summary.keep),
        ("Recommended: Review", summary.review),
        ("Recommended: Candidate for release", summary.candidate_release),
        ("Recommended: Do not release", summary.do_not_release),
        ("Recommended: Unknown", summary.unknown),
        ("Possible email dependencies", summary.email_dependency),
        ("Likely defensive / fraud-prevention (High)", summary.defensive_high),
    ]
    row = 3
    ws.cell(row=row, column=1, value="Metric").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Count").font = LABEL_FONT
    for cell in (ws.cell(row=row, column=1), ws.cell(row=row, column=2)):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1
    for label, value in metrics:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        row += 1

    # Patterns.
    row += 1
    ws.cell(row=row, column=1, value="Patterns").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value="Pattern group").font = LABEL_FONT
    ws.cell(row=row, column=2, value="Count").font = LABEL_FONT
    ws.cell(row=row, column=3, value="Examples").font = LABEL_FONT
    for cell in (ws.cell(row=row, column=1), ws.cell(row=row, column=2),
                 ws.cell(row=row, column=3)):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1
    for name, examples in summary.patterns.items():
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=len(examples))
        ws.cell(row=row, column=3, value=", ".join(examples[:12]))
        row += 1

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 90


def write_xlsx(rows: list[ClassifiedDomain], summary: Summary,
               path: str | Path) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_inventory_sheet(wb.active, rows)
    _write_summary_sheet(wb.create_sheet("Summary"), summary)
    wb.save(path)
    log.info("wrote xlsx: %s (%d rows)", path, len(rows))
    return path
